import logging
import json
import re
from datetime import datetime
from pathlib import Path

from utils import get_config

logger = logging.getLogger(__name__)

REPORTS_DIR = Path(get_config(
    'RADAR_REPORTS_DIR',
    str(Path(__file__).resolve().parent.parent / 'data' / 'reports'),
))

XLSX_HEADERS = [
    ('#', 4),
    ('Video ID', 14),
    ('URL', 50),
    ('Author', 20),
    ('Caption', 60),
    ('Source Type', 14),
    ('Source Value', 20),
    ('Views', 12),
    ('Likes', 12),
    ('Comments', 12),
    ('Shares', 12),
    ('Engagement Rate', 16),
    ('Comment Density', 16),
    ('Viral Score', 14),
    ('Subscriber Potential', 20),
    ('Final Score', 12),
    ('AI Analysis', 80),
]


def generate_report(
    stats: dict,
    top_videos: list[dict],
    ai_analyses: list[dict] | None = None,
    playlists: list[dict] | None = None,
) -> str:
    lines = [
        '# Nura TikTok Viral Screening Report',
        '',
        f'**Дата:** {stats.get("date", datetime.now().strftime("%Y-%m-%d %H:%M"))}',
        f'**Run ID:** {stats.get("run_id", "N/A")}',
        '',
        '## Статистика',
        '',
        f'- Источников обработано: {stats.get("sources_processed", 0)}',
        f'- Роликов найдено: {stats.get("videos_found", 0)}',
        f'- Новых роликов: {stats.get("new_videos", 0)}',
        f'- Роликов в отчёте: {len(top_videos)}',
        f'- Min просмотров: {stats.get("min_views", 10000)}',
        '',
        '### Легенда метрик',
        '',
        '| Метрика | Что показывает | Когда важна |',
        '|---|---|---|',
        '| **Final Score** | Вирусный потенциал (качество вовлечения) | Выбрать контент для адаптации под Nura |',
        '| **Subscriber Potential** | Способность набирать подписчиков (охват × вовлечение) | Выбрать ролик для посева / рекламы |',
        '| **Engagement Rate** | % зрителей, которые совершили действие | Оценить качество контента |',
        '| **Viral Score** | Во сколько раз просмотры превышают подписчиков | Найти видео, ушедшее в рекомендации |',
        '',
        '## Топ роликов',
        '',
    ]

    lines[5:5] = ['## Run Summary', '', f'- Mode: {stats.get("mode", "N/A")}', f'- Start/end: {stats.get("started_at", "N/A")} / {stats.get("completed_at", "N/A")}', '']
    lines += ['', '## Source Coverage', '', '| # | Source | Status | Raw | Unique | Added | Duplicates | Method |', '|---:|---|---|---:|---:|---:|---:|---|']
    for attempt in stats.get('source_attempts', []):
        lines.append(f'| {attempt.get("ordinal")} | {attempt.get("source_type")}/{attempt.get("source_value")} | {attempt.get("status")} | {attempt.get("raw_items_received", 0)} | {attempt.get("unique_within_source", 0)} | {attempt.get("unique_added_to_run", 0)} | {attempt.get("duplicates_already_seen_in_run", 0)} | {attempt.get("collection_method")} |')
    lines += ['', '## Cross-source Overlap', '']
    overlaps = [p for p in stats.get('provenance', []) if p.get('repeat_discoveries')]
    lines += [f'- {p.get("video_id")}: {len(p.get("matched_sources", []))} sources' for p in overlaps] or ['- No overlap observed.']
    buckets = {key: 0 for key in ('emerging', 'current', 'recent_evergreen', 'historical_evergreen', 'unknown')}
    for video in top_videos: buckets[video.get('freshness_bucket', 'unknown')] += 1
    lines += ['', '## Freshness Summary', ''] + [f'- {key}: {value}' for key, value in buckets.items()]
    lines += ['', '## Current Trend Candidates', ''] + [f'- {v.get("url")} ({v.get("freshness_bucket")})' for v in top_videos if v.get('freshness_bucket') in ('emerging', 'current')] or ['- None.']
    lines += ['', '## Evergreen Format References', ''] + [f'- {v.get("url")} ({v.get("freshness_bucket")})' for v in top_videos if v.get('freshness_bucket') in ('recent_evergreen', 'historical_evergreen')] or ['- None.']
    warnings = []
    if buckets['unknown']:
        warnings.append(f'missing published_at: {buckets["unknown"]} video(s)')
    for attempt in stats.get('source_attempts', []):
        if attempt.get('status') not in ('success', 'empty'):
            warnings.append(
                f'{attempt.get("source_type")}/{attempt.get("source_value")}: '
                f'{attempt.get("status")}'
            )
        if attempt.get('raw_items_received', 0) > attempt.get('unique_within_source', 0) * 3:
            warnings.append(
                f'{attempt.get("source_type")}/{attempt.get("source_value")}: duplicate-heavy or limit-truncated'
            )
    lines += ['', '## Data Quality Warnings', '']
    lines += [f'- {warning}' for warning in warnings] or ['- None.']

    analysis_map = {}
    if ai_analyses:
        for a in ai_analyses:
            analysis_map[a.get('video_id')] = a

    for i, v in enumerate(top_videos):
        caption = (v.get('caption') or 'Без описания')[:80]
        lines.append(f'### {i + 1}. {caption}')
        lines.append('')
        lines.append(f'- **Автор:** @{v.get("author_username", "N/A")}')
        lines.append(f'- **URL:** {v.get("url", "#")}')
        lines.append(f'- **Источник:** {v.get("source_type", "N/A")} / {v.get("source_value", "N/A")}')
        lines.append(f'- **Просмотры:** {_fmt(v.get("views"))}')
        lines.append(f'- **Лайки:** {_fmt(v.get("likes"))}')
        lines.append(f'- **Комментарии:** {_fmt(v.get("comments"))}')
        lines.append(f'- **Репосты:** {_fmt(v.get("shares"))}')
        lines.append(f'- **Engagement Rate:** {v.get("engagement_rate", 0):.2%}')
        lines.append(f'- **Comment Density:** {v.get("comment_density", 0):.4%}')
        if v.get('viral_score') is not None:
            lines.append(f'- **Viral Score:** {v.get("viral_score", 0):.1f}x')
        lines.append(f'- **Final Score:** {v.get("final_score", 0)}')
        lines.append(f'- **Freshness:** {v.get("freshness_bucket", "unknown")} ({v.get("published_at") or "unknown"})')
        lines.append(f'- **Score breakdown:** {v.get("score_breakdown", {})}')
        lines.append(f'- **Subscriber Potential:** {v.get("subscriber_potential", 0)}/10')
        lines.append('')

        video_id = v.get('video_id')
        if video_id and video_id in analysis_map:
            analysis = analysis_map[video_id].get('ai_analysis')
            if analysis:
                lines.append('#### AI-анализ:')
                lines.append('')
                lines.append(analysis)
                lines.append('')

        lines.append('---')
        lines.append('')

    if playlists:
        lines.append('## Найденные плейлисты')
        lines.append('')
        for pl in playlists[:10]:
            pl_caption = (pl.get('caption') or '[Playlist]')[:80]
            lines.append(
                f'- [{pl_caption}]({pl.get("url", "#")}) — '
                f'@{pl.get("author_username", "N/A")} '
                f'({pl.get("source_type", "")} / {pl.get("source_value", "")})'
            )
        lines.append('')

    return '\n'.join(lines)


def save_ai_analyses(top_videos: list[dict], analysis_map: dict | None = None) -> str | None:
    if not analysis_map:
        return None
    dir_path = REPORTS_DIR.parent / 'ai_analysis'
    dir_path.mkdir(parents=True, exist_ok=True)
    saved = 0
    for i, v in enumerate(top_videos[:10], 1):
        video_id = v.get('video_id', 'unknown')
        analysis = ''
        if video_id in analysis_map:
            analysis = analysis_map[video_id].get('ai_analysis', '') or ''
        caption = (v.get('caption') or 'Без описания')[:40]
        safe_caption = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', caption).strip('._ ')[:40] or 'nocaption'
        safe_name = f'{i:02d}_{video_id}_{safe_caption}'[:100]
        lines = [
            f'#{i} — {caption}',
            f'Video ID: {video_id}',
            f'URL: {v.get("url", "")}',
            f'Author: @{v.get("author_username", "N/A")}',
            f'Views: {_fmt(v.get("views"))}',
            f'Likes: {_fmt(v.get("likes"))}',
            f'Score: {v.get("final_score", 0)}',
            '',
            '─' * 60,
            '',
            analysis or 'AI-анализ не выполнен',
            '',
        ]
        path = dir_path / f'{safe_name}.txt'
        with open(path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        saved += 1
    logger.info(f'AI analyses saved: {saved} files to {dir_path}')
    return str(dir_path)


def _fmt(value) -> str:
    if value is None:
        return 'N/A'
    try:
        return f'{int(value):,}'
    except (ValueError, TypeError):
        return str(value)


def save_report(markdown: str) -> str:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime('%Y-%m-%d_%H-%M')
    path = REPORTS_DIR / f'report_{ts}.md'
    with open(path, 'w', encoding='utf-8') as f:
        f.write(markdown)
    logger.info(f"Report saved: {path}")
    return str(path)


def save_xlsx(top_videos: list[dict], analysis_map: dict | None = None, stats: dict | None = None) -> str | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        logger.warning("openpyxl not installed. Run: pip install openpyxl")
        return None

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime('%Y-%m-%d_%H-%M')
    path = REPORTS_DIR / f'report_{ts}.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Videos'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2F2F2F', end_color='2F2F2F', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx, (name, width) in enumerate(XLSX_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = max(
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width or 8, width
        )

    for row_idx, v in enumerate(top_videos, 2):
        video_id = v.get('video_id', '')
        analysis = ''
        if analysis_map and video_id in analysis_map:
            analysis = analysis_map[video_id].get('ai_analysis', '')

        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=video_id)
        ws.cell(row=row_idx, column=3, value=v.get('url', ''))
        ws.cell(row=row_idx, column=4, value=v.get('author_username', ''))
        ws.cell(row=row_idx, column=5, value=(v.get('caption') or '')[:200])
        ws.cell(row=row_idx, column=6, value=v.get('source_type', ''))
        ws.cell(row=row_idx, column=7, value=v.get('source_value', ''))
        ws.cell(row=row_idx, column=8, value=v.get('views'))
        ws.cell(row=row_idx, column=9, value=v.get('likes'))
        ws.cell(row=row_idx, column=10, value=v.get('comments'))
        ws.cell(row=row_idx, column=11, value=v.get('shares'))
        ws.cell(row=row_idx, column=12, value=v.get('engagement_rate'))
        ws.cell(row=row_idx, column=13, value=v.get('comment_density'))
        ws.cell(row=row_idx, column=14, value=v.get('viral_score'))
        ws.cell(row=row_idx, column=15, value=v.get('subscriber_potential'))
        ws.cell(row=row_idx, column=16, value=v.get('final_score'))
        ws.cell(row=row_idx, column=17, value='AI analysis:\n' + (analysis[:10000] if analysis else ''))

    stats = stats or {}
    for title, headers, rows in (
        ('Source Coverage', ['Ordinal', 'Source', 'Status', 'Raw', 'Unique', 'Added', 'Duplicates'], [[a.get('ordinal'), f'{a.get("source_type")}/{a.get("source_value")}', a.get('status'), a.get('raw_items_received'), a.get('unique_within_source'), a.get('unique_added_to_run'), a.get('duplicates_already_seen_in_run')] for a in stats.get('source_attempts', [])]),
        ('Provenance Matches', ['Video ID', 'URL', 'Matches', 'Repeats', 'New to DB'], [[p.get('video_id'), p.get('canonical_url'), len(p.get('matched_sources', [])), p.get('repeat_discoveries'), p.get('new_to_database')] for p in stats.get('provenance', [])]),
        ('Run Summary', ['Field', 'Value'], [[
            k,
            json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v,
        ] for k, v in stats.items() if k not in ('source_attempts', 'provenance')]),
    ):
        extra = wb.create_sheet(title); extra.append(headers)
        for row in rows: extra.append(row)
    wb.save(str(path))
    logger.info(f"XLSX report saved: {path}")
    return str(path)


SCENARIO_HEADERS = [
    ('A. ID видео', 14),
    ('B. Формат', 20),
    ('C. Цель', 16),
    ('D. Локация / Фон', 30),
    ('E. Крючок (Текст на экране)', 40),
    ('F. Полный скрипт озвучки (Речь Нуры)', 60),
    ('G. Стоковые вставки', 40),
    ('H. Динамика камеры и эффекты', 30),
    ('I. Стиль и позиция субтитров', 30),
    ('J. Директива для Карусели (Instagram)', 30),
    ('K. Переход', 18),
    ('L. Темп (BPM)', 14),
    ('M. Тональность', 20),
]


def _parse_scenario_block(text: str) -> list[dict]:
    scenarios = []
    blocks = re.split(r'={3,}\s*СЦЕНАРИЙ\s*\d+\s*={3,}', text)
    for block in blocks:
        block = block.strip()
        if not block or len(block) < 20:
            continue
        fields = {
            'Формат': '',
            'Цель': '',
            'Локация': '',
            'Крючок': '',
            'Скрипт озвучки': '',
            'Стоковые вставки': '',
            'Камера и эффекты': '',
            'Субтитры': '',
            'Карусель': '',
            'Переход': '',
            'Темп': '',
            'Тональность': '',
        }
        for key in fields:
            m = re.search(rf'{re.escape(key)}\s*:\s*(.*?)(?:\n(?:[A-ZА-Я][A-ZА-Яa-zа-я\s]+):|\Z)', block, re.DOTALL)
            if m:
                fields[key] = m.group(1).strip()
        scenarios.append(fields)
    return scenarios


def _parse_carousel_block(text: str) -> str | None:
    m = re.search(
        r'={3,}\s*КАРУСЕЛЬ\s*={3,}\s*\n(.*?)\n[\s]*={3,}\s*/КАРУСЕЛЬ\s*={3,}',
        text, re.DOTALL | re.IGNORECASE,
    )
    if not m:
        m = re.search(
            r'={3,}\s*КАРУСЕЛЬ\s*={3,}\s*\n(.*?)(?:\n\s*={3,}|\Z)',
            text, re.DOTALL | re.IGNORECASE,
        )
    if m:
        content = m.group(1).strip()
        if content and 'не применимо' not in content.lower()[:20]:
            return content
    return None


def save_carousel_texts(
    top_videos: list[dict], analysis_map: dict | None = None
) -> str | None:
    if not analysis_map:
        return None
    dir_path = REPORTS_DIR.parent / 'carousel_texts'
    dir_path.mkdir(parents=True, exist_ok=True)
    saved = 0
    for i, v in enumerate(top_videos[:10], 1):
        video_id = v.get('video_id', 'unknown')
        analysis = ''
        if video_id in analysis_map:
            analysis = analysis_map[video_id].get('ai_analysis', '') or ''
        carousel_text = _parse_carousel_block(analysis)
        if not carousel_text:
            continue
        caption = (v.get('caption') or 'Без описания')[:40]
        safe_caption = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', caption).strip('._ ')[:40] or 'nocaption'
        safe_name = f'{i:02d}_{video_id}_{safe_caption}'[:100]
        path = dir_path / f'{safe_name}.carousel.txt'
        with open(path, 'w', encoding='utf-8') as f:
            f.write(carousel_text)
        saved += 1
    logger.info(f'Carousel texts saved: {saved} files to {dir_path}')
    return str(dir_path)


def save_scenarios_xlsx(top_videos: list[dict], analysis_map: dict | None = None) -> str | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        logger.warning("openpyxl not installed. Run: pip install openpyxl")
        return None

    if not analysis_map:
        return None

    dir_path = REPORTS_DIR.parent / 'scenarios'
    dir_path.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime('%Y-%m-%d_%H-%M')
    path = dir_path / f'scenarios_{ts}.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Сценарии Nura'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(vertical='top', wrap_text=True)

    for col_idx, (name, width) in enumerate(SCENARIO_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[chr(64 + col_idx)].width = width

    row_idx = 2
    for i, v in enumerate(top_videos[:10], 1):
        video_id = v.get('video_id', 'unknown')
        analysis = ''
        if video_id in analysis_map:
            analysis = analysis_map[video_id].get('ai_analysis', '') or ''
        scenarios = _parse_scenario_block(analysis)

        if not scenarios:
            ws.cell(row=row_idx, column=1, value=f'{video_id}_1')
            ws.cell(row=row_idx, column=6, value=analysis[:10000])
            ws.cell(row=row_idx, column=1).alignment = cell_align
            ws.cell(row=row_idx, column=6).alignment = cell_align
            row_idx += 1
            continue

        for si, sc in enumerate(scenarios, 1):
            ws.cell(row=row_idx, column=1, value=f'{video_id}_{si}')
            ws.cell(row=row_idx, column=2, value=sc.get('Формат', ''))
            ws.cell(row=row_idx, column=3, value=sc.get('Цель', ''))
            ws.cell(row=row_idx, column=4, value=sc.get('Локация', ''))
            ws.cell(row=row_idx, column=5, value=sc.get('Крючок', ''))
            ws.cell(row=row_idx, column=6, value=sc.get('Скрипт озвучки', ''))
            ws.cell(row=row_idx, column=7, value=sc.get('Стоковые вставки', ''))
            ws.cell(row=row_idx, column=8, value=sc.get('Камера и эффекты', ''))
            ws.cell(row=row_idx, column=9, value=sc.get('Субтитры', ''))
            ws.cell(row=row_idx, column=10, value=sc.get('Карусель', ''))
            ws.cell(row=row_idx, column=11, value=sc.get('Переход', ''))
            ws.cell(row=row_idx, column=12, value=sc.get('Темп', ''))
            ws.cell(row=row_idx, column=13, value=sc.get('Тональность', ''))
            for col in range(1, 14):
                ws.cell(row=row_idx, column=col).alignment = cell_align
            row_idx += 1

    wb.save(str(path))
    logger.info(f'Scenarios saved: {path} ({row_idx - 2} rows)')
    return str(path)
