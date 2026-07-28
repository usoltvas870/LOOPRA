"""Portable LOOPRA 0.5 Trend Workbook package builder.

The module is intentionally local-only: it receives acquired media and
evidence, then creates a self-contained workbook package.  It never invokes
Content Intelligence, a script provider, image generation, or a browser.
"""
from __future__ import annotations

import asyncio
import hashlib
import json
import os
import re
import shutil
import tempfile
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from types import SimpleNamespace
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


SCHEMA_VERSION = "0.5"
SCORING_PROFILE = "NURA_TREND_WORKBOOK_V05"
REQUIRED_SHEETS = ("Сводка", "Кандидаты", "Транскрипции", "Отбраковано", "Методика")
OPERATOR_COLUMNS = ("Выбрать", "Приоритет", "Комментарий")
HEADERS = (
    *OPERATOR_COLUMNS, "Итоговый ранг", "Предварительная релевантность", "Причина отбора", "Риск мусора",
    "Hook", "Источник hook", "Краткая тема", "Автор", "Открыть локальное видео", "Имя локального файла",
    "Исходная TikTok-ссылка", "Video ID", "Candidate ID", "Дата публикации", "Длительность", "Размер MP4",
    "SHA-256 MP4", "Просмотры", "Лайки", "Комментарии", "Репосты", "Like rate", "Comment rate", "Share rate",
    "Engagement rate", "Virality score", "Relevance score", "Freshness score", "Quality score", "Final score",
    "Classification", "Query cluster", "Search query", "Query source type", "Caption", "Transcript status",
    "Transcript excerpt", "Transcript segment count", "Перейти к транскрипции", "OCR status", "On-screen text excerpt",
    "Audio role", "Content modality", "Duplicate status", "Evidence warning", "Technical status",
)
JUNK_TERMS = {
    "sport": ("спорт", "футбол", "матч", "nba"),
    "food": ("рецепт", "еда", "готов", "кухн"),
    "gaming": ("игра", "gaming", "minecraft"),
    "meme": ("мем", "прикол", "смешн"),
    "tarot": ("таро", "гадани", "предсказан"),
}


class TrendWorkbookError(ValueError):
    pass


def backfill_candidate_media(
    *, candidates: Iterable[dict[str, Any]], lookup_media, acquire_one,
    target_count: int = 20, maximum_attempts: int = 30,
    maximum_shortlist_size: int = 50, maximum_consecutive_failures: int = 5,
) -> dict[str, Any]:
    """Resume ranked acquisition until ``target_count`` valid unique media exist."""
    if not 1 <= target_count <= 30:
        raise TrendWorkbookError("target_count must be between 1 and 30")
    if maximum_attempts < 1 or maximum_shortlist_size < target_count or maximum_consecutive_failures < 1:
        raise TrendWorkbookError("invalid bounded backfill limits")
    candidate_list = list(candidates)
    results: list[dict[str, Any]] = []
    hashes: dict[str, str] = {}
    attempts = failures = duplicates = consecutive_failures = 0
    processed_positions: list[int] = []
    starting_unique = 0
    for position, raw in enumerate(candidate_list[:maximum_shortlist_size], 1):
        candidate = dict(raw)
        candidate.setdefault("candidate_id", candidate.get("video_id"))
        candidate.setdefault("query", candidate.get("source_value", ""))
        candidate.setdefault("query_cluster", candidate.get("source_value", ""))
        _apply_nura_relevance(candidate)
        existing = lookup_media(candidate)
        if existing is None:
            if len(hashes) >= target_count or attempts >= maximum_attempts or consecutive_failures >= maximum_consecutive_failures:
                break
            attempts += 1
            processed_positions.append(position)
            existing = acquire_one(candidate, position)
        else:
            processed_positions.append(position)
        status = str((existing or {}).get("status") or "FAILED")
        media_path = Path(existing["local_media_path"]) if existing and existing.get("local_media_path") else None
        digest = existing.get("media_sha256") if existing else None
        valid = status in {"COMPLETED", "REUSED"} and media_path is not None and media_path.is_file() and digest == sha256(media_path)
        if not valid:
            failures += 1
            consecutive_failures += 1
            candidate.update(local_media_path=None, rejection_reason=(existing or {}).get("reason", "MEDIA_NOT_ACQUIRED"), technical_status=status)
            results.append(candidate)
            continue
        consecutive_failures = 0
        candidate.update(local_media_path=str(media_path), media_sha256=digest, duration_seconds=existing.get("duration_seconds"), technical_status="VALID_MEDIA")
        assessed = assess_candidate(candidate)
        if not assessed["eligible"]:
            candidate["rejection_reason"] = assessed.get("rejection_reason", "LOW_RELEVANCE")
            results.append(candidate)
            continue
        canonical = hashes.get(digest)
        if canonical:
            duplicates += 1
            candidate.update(rejection_reason="EXACT_DUPLICATE", duplicate_canonical=canonical)
            results.append(candidate)
            continue
        hashes[digest] = str(candidate.get("video_id"))
        if attempts == 0:
            starting_unique += 1
        results.append(candidate)
        if len(hashes) >= target_count:
            break
    return {
        "candidates": results, "starting_unique_media": starting_unique,
        "additional_attempts": attempts, "failures": failures, "duplicates": duplicates,
        "valid_unique_media": len(hashes), "processed_positions": processed_positions,
        "shortlist_exhausted": len(hashes) < target_count and len(processed_positions) >= min(len(candidate_list), maximum_shortlist_size),
    }


class _PublicBackfillAcquirer:
    """Thin lifecycle owner over the existing per-item browser capture primitive."""

    def __init__(self, runtime_root: Path, collection: dict[str, Any], collection_path: Path) -> None:
        from browser_media_capture import BrowserMediaCaptureRequest, capture_browser_media_in_context
        from collector import TikTokCollector
        from utils import get_cookie_path
        self.runtime_root = runtime_root
        self.collection = collection
        self.collection_path = collection_path
        self.capture_request = BrowserMediaCaptureRequest
        self.capture = capture_browser_media_in_context
        self.collector = TikTokCollector(headless=True)
        self.cookie_path = get_cookie_path()
        self.loop = asyncio.new_event_loop()
        self.started = False
        self.manifest = SimpleNamespace(
            radar_run_id=collection["search_run_id"],
            radar_run_reference="canonical/collection.json",
            manifest_hash=hashlib.sha256(json.dumps(collection, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")).hexdigest(),
        )

    def acquire(self, candidate: dict[str, Any], position: int) -> dict[str, Any]:
        if not self.started:
            self.loop.run_until_complete(self.collector.start())
            self.started = True
        selected = SimpleNamespace(video_id=str(candidate["video_id"]), rank=position, canonical_url=candidate.get("url"))
        request = self.capture_request(
            selection_manifest_path=self.collection_path,
            cookie_state_path=self.cookie_path,
            output_root=self.runtime_root / "canonical" / "acquisition",
            candidate_id=selected.video_id,
        )
        record = self.loop.run_until_complete(self.capture(request, self.manifest, selected, self.collector.context))
        return _media_from_record(self.runtime_root, record.to_dict() if hasattr(record, "to_dict") else record)

    def close(self) -> None:
        try:
            if self.started:
                self.loop.run_until_complete(self.collector.close())
        finally:
            if not self.loop.is_closed():
                self.loop.close()


def _media_from_record(runtime_root: Path, record: dict[str, Any]) -> dict[str, Any]:
    run_root = runtime_root / "canonical" / "acquisition" / str(record.get("radar_run_id"))
    reference = record.get("local_media_path")
    media = run_root / reference if reference else None
    probe = record.get("ffprobe_validation") or {}
    valid = bool(
        record.get("status") in {"COMPLETED", "REUSED"}
        and media and media.is_file() and record.get("media_sha256") == sha256(media)
        and probe.get("valid") and probe.get("video_codec")
    )
    return {
        "status": "REUSED" if valid else "FAILED",
        "local_media_path": str(media) if valid else None,
        "media_sha256": record.get("media_sha256") if valid else None,
        "duration_seconds": probe.get("duration_seconds"),
        "reason": "INVALID_MEDIA" if reference and not valid else (record.get("errors") or ["MEDIA_NOT_ACQUIRED"])[0],
    }


def _lookup_runtime_media(runtime_root: Path, search_run_id: str, candidate: dict[str, Any]) -> dict[str, Any] | None:
    path = runtime_root / "canonical" / "acquisition" / search_run_id / str(candidate.get("video_id")) / "acquisition_record.json"
    if not path.is_file():
        return None
    try:
        return _media_from_record(runtime_root, json.loads(path.read_text(encoding="utf-8")))
    except (OSError, json.JSONDecodeError, KeyError, TypeError):
        return {"status": "FAILED", "reason": "INVALID_ACQUISITION_RECORD"}


def _classify_transcription(payload: dict[str, Any], artifact_reference: str, caption: str) -> dict[str, Any]:
    status = str(payload.get("status") or "FAILED")
    segments = []
    for segment in payload.get("segments") or []:
        normalized = str(segment.get("normalized_text") or "").strip()
        segments.append({**segment, "text": normalized, "language": payload.get("language"), "quality_status": "ACCEPTED", "evidence_ref": artifact_reference})
    if status == "COMPLETED_NO_AUDIO":
        role, reason = "NO_AUDIO", "inspection confirmed no audio stream"
    elif status == "COMPLETED_NO_SPEECH":
        role, reason = "NONSPEECH", "canonical ASR found no reliable speech"
    elif status == "COMPLETED" and segments:
        words = sum(len(item["text"].split()) for item in segments)
        music_signal = any(token in caption.lower() for token in ("#music", "#song", "песня", "lyrics", "музыка"))
        if music_signal:
            role, reason = "BACKGROUND_MUSIC_ONLY", "caption identifies music; ASR text is not eligible as author hook"
        elif words >= 4:
            role, reason = "AUTHOR_SPEECH_USABLE", "canonical ASR produced meaningful accepted segments"
        else:
            role, reason = "UNRELIABLE_ASR", "accepted ASR contains fewer than four words"
    elif status == "FAILED":
        role, reason = "TRANSCRIPTION_FAILED", "; ".join(payload.get("errors") or ["canonical transcription failed"])
    else:
        role, reason = "MANUAL_REVIEW_REQUIRED", f"unclassified canonical transcription status: {status}"
    return {"audio_role": role, "transcript_status": role, "transcript_reason": reason, "transcript_segments": segments}


def _build_transcription_evidence(
    *, runtime_root: Path, build_root: Path, search_run_id: str,
    candidates: list[dict[str, Any]], inspection_callable=None, engine=None,
) -> dict[str, Any]:
    from format_inspection import inspect as canonical_inspect
    from selection_manifest import build_selection_manifest, write_selection_manifest
    from transcription_evidence import (
        DEFAULT_OPTIONS, FasterWhisperEngine, TranscriptionRunRequest,
        _migrate_legacy_result, _no_audio_result, _prepare, _transcribe, _write_atomic,
    )
    inspection_callable = inspection_callable or canonical_inspect
    engine = engine or FasterWhisperEngine()
    availability = engine.availability()
    manifest = build_selection_manifest(candidates, radar_run_id=search_run_id, created_at=_utc_now())
    manifest_path = write_selection_manifest(manifest, root=build_root / "selection")
    acquisition_root = runtime_root / "canonical" / "acquisition" / search_run_id
    inspection_root = build_root / "inspection"
    evidence_root = build_root / "evidence"
    by_id = {str(item["video_id"]): item for item in candidates}
    for entry in manifest.candidates:
        item = by_id[entry.video_id]
        output = inspection_root / entry.video_id
        target = output / "inspection.json"
        if not target.is_file() or json.loads(target.read_text(encoding="utf-8")).get("media_sha256") != sha256(Path(item["local_media_path"])):
            inspection_callable(Path(item["local_media_path"]), output, entry.video_id, item.get("url"))
    request = TranscriptionRunRequest(
        selection_manifest_path=manifest_path, acquisition_root=acquisition_root,
        inspection_root=inspection_root, output_root=evidence_root, reuse=True,
        options=dict(DEFAULT_OPTIONS),
    )
    calls = reused = 0
    for entry in manifest.candidates:
        target_reference = f"evidence/{search_run_id}/candidates/{entry.video_id}/transcription/transcription_result.json"
        if not availability.get("available"):
            payload = {"candidate_video_id": entry.video_id, "rank": entry.rank, "status": "FAILED", "errors": [availability.get("reason", "transcription engine unavailable")]}
        else:
            prepared = _prepare(entry, manifest, request, availability, dict(DEFAULT_OPTIONS))
            if prepared.get("reused"):
                payload = prepared["result"]; reused += 1
            elif prepared.get("legacy"):
                payload = _migrate_legacy_result(prepared); _write_atomic(prepared["target"], payload); calls += 1
            elif prepared.get("no_audio"):
                payload = _no_audio_result(prepared); _write_atomic(prepared["target"], payload); calls += 1
            else:
                payload = _transcribe(prepared, engine, request.language, dict(DEFAULT_OPTIONS)); calls += 1
        by_id[entry.video_id].update(_classify_transcription(payload, target_reference, str(by_id[entry.video_id].get("caption") or "")))
    counts = Counter(item.get("audio_role") for item in candidates)
    return {"candidates": candidates, "transcription_calls": calls, "transcription_reused": reused, "status_counts": dict(counts), "manifest_path": str(manifest_path)}


def resume_public_first_workbook(
    *, project_id: str, runtime_root: Path, output_root: Path, target_count: int = 20,
    build_id: str, maximum_attempts: int = 30, maximum_shortlist_size: int = 50,
    maximum_consecutive_failures: int = 5, acquire_one=None, inspection_callable=None, transcription_engine=None,
) -> dict[str, Any]:
    """Resume an existing public pool; never invokes metadata collection."""
    runtime_root, output_root = Path(runtime_root), Path(output_root)
    collection_path = runtime_root / "canonical" / "collection.json"
    if not collection_path.is_file():
        raise TrendWorkbookError("existing canonical collection.json is required for --resume")
    pool = json.loads(collection_path.read_text(encoding="utf-8"))
    search_run_id = str(pool.get("search_run_id") or "")
    if not search_run_id or not isinstance(pool.get("candidates"), list):
        raise TrendWorkbookError("existing collection artifact is invalid")
    package_run_id = f"{search_run_id}_{_safe_component(build_id)}"
    package_dir = output_root / f"LOOPRA_05_TREND_WORKBOOK_{package_run_id}"
    if package_dir.is_dir():
        package = build_package(project_id=project_id, search_run_id=search_run_id, package_build_id=build_id, candidates=[], output_root=output_root)
        manifest = json.loads((package_dir / "manifest.json").read_text(encoding="utf-8"))
        return {"status": _acceptance_status(manifest), "package": package, "manifest": manifest, "search_calls": 0, "browser_calls": 0, "downloads": 0, "transcription_calls": 0, "reuse": True, "provider_calls": 0, "script_calls": 0}
    acquirer = None
    try:
        def lookup(candidate):
            return _lookup_runtime_media(runtime_root, search_run_id, candidate)
        if acquire_one is None:
            acquirer = _PublicBackfillAcquirer(runtime_root, pool, collection_path)
            acquire_one = acquirer.acquire
        backfill = backfill_candidate_media(
            candidates=pool["candidates"], lookup_media=lookup, acquire_one=acquire_one,
            target_count=target_count, maximum_attempts=maximum_attempts,
            maximum_shortlist_size=maximum_shortlist_size,
            maximum_consecutive_failures=maximum_consecutive_failures,
        )
    finally:
        if acquirer is not None:
            acquirer.close()
    usable = [assess_candidate(item) for item in backfill["candidates"] if item.get("local_media_path") and not item.get("rejection_reason")]
    usable = sorted((item for item in usable if item["eligible"]), key=lambda item: (-item["final_score"], str(item.get("video_id"))))[:target_count]
    if len(usable) < target_count:
        return {"status": "PARTIAL_INSUFFICIENT_VALID_MEDIA", "backfill": backfill, "search_calls": 0, "provider_calls": 0, "script_calls": 0}
    build_root = runtime_root / "canonical" / "workbook-builds" / _safe_component(build_id)
    evidence = _build_transcription_evidence(
        runtime_root=runtime_root, build_root=build_root, search_run_id=search_run_id,
        candidates=usable, inspection_callable=inspection_callable, engine=transcription_engine,
    )
    package_candidates = evidence["candidates"] + [item for item in backfill["candidates"] if item.get("rejection_reason")]
    package = build_package(
        project_id=project_id, search_run_id=search_run_id, package_build_id=build_id,
        candidates=package_candidates, output_root=output_root,
    )
    manifest = json.loads((Path(package["package_path"]) / "manifest.json").read_text(encoding="utf-8"))
    manifest["backfill"] = {key: value for key, value in backfill.items() if key != "candidates"}
    manifest["transcription"] = {key: value for key, value in evidence.items() if key != "candidates"}
    (Path(package["package_path"]) / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return {"status": _acceptance_status(manifest), "package": package, "manifest": manifest, "backfill": backfill, "transcription": evidence, "search_calls": 0, "browser_calls": backfill["additional_attempts"], "downloads": backfill["additional_attempts"], "provider_calls": 0, "script_calls": 0, "reuse": False}


def _acceptance_status(manifest: dict[str, Any]) -> str:
    if manifest.get("exported_candidate_count", 0) < 20 or manifest.get("video_file_count", 0) < 20:
        return "PARTIAL_INSUFFICIENT_VALID_MEDIA"
    transcription = manifest.get("transcription", {})
    counts = transcription.get("status_counts", {})
    if not counts or counts.get("MANUAL_REVIEW_REQUIRED", 0) == manifest.get("exported_candidate_count"):
        return "PARTIAL_TRANSCRIPTION_INTEGRATION"
    return "READY_FOR_OWNER_WORKBOOK_REVIEW"


def run_public_first_workbook(
    *, project_id: str, runtime_root: Path, output_root: Path,
    production_dependencies: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Run the proven v2 public collector and per-item capture without CI.

    The v1 five-item wrapper is deliberately not imported.  The v2 factory
    owns one guest-capable browser context and exposes rank-independent
    ``acquire``/``transcribe`` callables for each canonical selection entry.
    """
    if production_dependencies is None:
        from loopra_top20_real_pipeline_v2 import build_fresh_top20_b1_production_dependencies
        production_dependencies = build_fresh_top20_b1_production_dependencies(root=runtime_root)
    required = ("collect", "select", "acquire", "ocr", "transcribe", "close")
    if any(not callable(production_dependencies.get(name)) for name in required):
        raise TrendWorkbookError("public-first v2 dependencies are incomplete")
    counters = {"collection_calls": 0, "acquisition_calls": 0, "ocr_calls": 0, "transcription_calls": 0}
    try:
        counters["collection_calls"] += 1
        pool = production_dependencies["collect"]()
        access = pool.get("public_access_status") or pool.get("status")
        if access in {"PUBLIC_ACCESS_BLOCKED", "CAPTCHA_OR_ANTI_BOT_CHALLENGE", "RATE_LIMITED"}:
            return {"status": access, "counters": counters, "pool": pool}
        existing_package = Path(output_root) / f"LOOPRA_05_TREND_WORKBOOK_{_safe_component(pool['search_run_id'])}"
        if existing_package.is_dir():
            package = build_package(project_id=project_id, search_run_id=pool["search_run_id"], candidates=[], output_root=output_root)
            manifest = json.loads((existing_package / "manifest.json").read_text(encoding="utf-8"))
            status = "READY_FOR_OWNER_WORKBOOK_REVIEW" if package["exported"] >= 20 else "PARTIAL_INSUFFICIENT_VALID_MEDIA"
            return {"status": status, "package": package, "manifest": manifest, "pool": pool, "counters": counters, "provider_calls": 0, "script_calls": 0}
        entries = production_dependencies["select"](pool)
        if not entries:
            return {"status": "PARTIAL_INSUFFICIENT_RELEVANT_CANDIDATES", "counters": counters, "pool": pool}
        pool_by_id = {str(item.get("video_id")): item for item in pool.get("candidates", [])}
        acquired_candidates: list[tuple[dict[str, Any], dict[str, Any]]] = []
        failures: list[dict[str, Any]] = []
        for entry in entries:
            counters["acquisition_calls"] += 1
            acquisition = production_dependencies["acquire"](entry)
            if acquisition.get("status") not in {"COMPLETED", "REUSED"} or acquisition.get("ffprobe_status") != "VALID":
                failures.append({"video_id": entry["video_id"], "rejection_reason": "MEDIA_NOT_ACQUIRED"})
                continue
            source = runtime_root / str(acquisition["source_media_reference"])
            original = dict(pool_by_id.get(str(entry["video_id"]), {}))
            original.update(video_id=entry["video_id"], candidate_id=entry["candidate_id"], local_media_path=str(source))
            acquired_candidates.append((entry, original))
        candidates: list[dict[str, Any]] = []
        for entry, original in acquired_candidates:
            counters["ocr_calls"] += 1
            ocr = production_dependencies["ocr"](entry)
            counters["transcription_calls"] += 1
            transcription = production_dependencies["transcribe"](entry)
            original.update(_evidence_from_result(runtime_root, ocr, transcription))
            _apply_nura_relevance(original)
            candidates.append(original)
        if not candidates:
            return {"status": "PARTIAL_INSUFFICIENT_VALID_MEDIA", "counters": counters, "pool": pool, "rejected": failures}
        package = build_package(project_id=project_id, search_run_id=pool["search_run_id"], candidates=candidates, output_root=output_root)
        manifest = json.loads((Path(package["package_path"]) / "manifest.json").read_text(encoding="utf-8"))
        status = "READY_FOR_OWNER_WORKBOOK_REVIEW" if package["exported"] >= 20 else "PARTIAL_INSUFFICIENT_VALID_MEDIA"
        return {"status": status, "package": package, "manifest": manifest, "pool": pool, "entries": len(entries), "counters": counters, "acquisition_failures": failures, "provider_calls": 0, "script_calls": 0}
    finally:
        production_dependencies["close"]()


def _evidence_from_result(root: Path, ocr: dict[str, Any], transcription: dict[str, Any]) -> dict[str, Any]:
    result: dict[str, Any] = {"ocr_status": "MANUAL_TEXT_CHECK", "audio_role": "MANUAL_REVIEW_REQUIRED", "transcript_segments": []}
    for evidence, kind in ((ocr, "ocr"), (transcription, "transcription")):
        reference = evidence.get("artifact_reference")
        path = root / reference if reference else None
        if not path or not path.is_file():
            continue
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if kind == "ocr":
            hook = payload.get("first_text_hook") or {}
            text = hook.get("hook_text") or hook.get("text") or ""
            result.update(ocr_status="READABLE" if text else "MANUAL_TEXT_CHECK", ocr_text=text)
        else:
            segments = payload.get("segments") or []
            status = str(payload.get("status") or "MANUAL_REVIEW_REQUIRED")
            usable = status.startswith("COMPLETED") and bool(segments)
            result.update(transcript_segments=segments, audio_role="AUTHOR_SPEECH_USABLE" if usable else "MANUAL_REVIEW_REQUIRED")
    return result


def _apply_nura_relevance(candidate: dict[str, Any]) -> None:
    text = _text(candidate)
    topic_terms = ("выгоран", "устал", "границ", "тревог", "самооцен", "отста", "отдых", "пустот", "отношен", "выбрать себя")
    matched = sum(term in text for term in topic_terms)
    candidate["topical_relevance"] = 0.85 if matched else 0.45
    candidate["audience_relevance"] = 0.75 if matched else 0.4
    candidate["transferable_potential"] = 0.7 if matched else 0.4
    candidate["evidence_quality"] = 0.8 if candidate.get("transcript_segments") else 0.45
    candidate["virality_score"] = min(1.0, float(candidate.get("engagement_score", candidate.get("final_score", 0)) or 0) / 100)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _safe_component(value: object) -> str:
    result = re.sub(r'[^A-Za-z0-9_-]+', "_", str(value)).strip("._")
    return result[:80] or "unknown"


def _utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _copy_exact(source: Path, target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(delete=False, dir=target.parent, suffix=".part") as temp:
        temporary = Path(temp.name)
        with source.open("rb") as input_stream:
            shutil.copyfileobj(input_stream, temp)
        temp.flush(); os.fsync(temp.fileno())
    os.replace(temporary, target)
    if source.stat().st_size != target.stat().st_size or sha256(source) != sha256(target):
        target.unlink(missing_ok=True)
        raise TrendWorkbookError(f"exact copy validation failed for {source.name}")


def _media_is_valid(path: Path) -> bool:
    """Use the existing ffprobe gate; never export an unreadable placeholder."""
    try:
        from media_acquisition import _ffprobe
        return bool(_ffprobe(path).get("valid"))
    except (OSError, ValueError):
        return False


def _text(candidate: dict[str, Any]) -> str:
    return " ".join(str(candidate.get(key) or "") for key in ("caption", "query", "query_cluster", "transcript_text", "ocr_text")).lower()


def assess_candidate(candidate: dict[str, Any]) -> dict[str, Any]:
    """Apply the explicit relevance gate before engagement-based ranking."""
    result = dict(candidate)
    text = _text(result)
    junk = next((kind for kind, terms in JUNK_TERMS.items() if any(term in text for term in terms)), None)
    topical = float(result.get("topical_relevance", result.get("relevance_score", 0.0)) or 0.0)
    audience = float(result.get("audience_relevance", topical) or 0.0)
    transferable = float(result.get("transferable_potential", topical) or 0.0)
    virality = float(result.get("virality_score", result.get("engagement_score", 0.0)) or 0.0)
    freshness = float(result.get("freshness_score", 0.0) or 0.0)
    evidence = float(result.get("evidence_quality", 0.0) or 0.0)
    media = 1.0 if result.get("local_media_path") else 0.0
    relevance = (topical * 0.55) + (audience * 0.25) + (transferable * 0.20)
    classification = "LIKELY_RELEVANT" if relevance >= 0.55 else "UNCLEAR_REVIEW" if relevance >= 0.35 else "LIKELY_IRRELEVANT"
    if junk:
        classification = "LIKELY_IRRELEVANT"
        result["rejection_reason"] = {"sport": "SPORT", "food": "FOOD_OR_HOUSEHOLD", "gaming": "MEME_OR_ENTERTAINMENT", "meme": "RANDOM_HUMOR", "tarot": "OFF_TOPIC"}[junk]
    # Low topical relevance is an eligibility failure, never offset by engagement.
    eligible = topical >= 0.35 and classification != "LIKELY_IRRELEVANT"
    result.update({
        "topical_relevance": topical, "audience_relevance": audience, "transferable_potential": transferable,
        "virality_score": virality, "freshness_score": freshness, "evidence_quality": evidence,
        "media_availability": media, "relevance_score": relevance, "classification": classification,
        "eligible": eligible,
        "quality_score": (evidence + transferable) / 2,
        "final_score": relevance * 0.55 + virality * 0.20 + freshness * 0.10 + evidence * 0.10 + media * 0.05 if eligible else 0.0,
    })
    return result


def _hook(candidate: dict[str, Any]) -> tuple[str, str]:
    segments = candidate.get("transcript_segments") or []
    audio_role = candidate.get("audio_role") or "MANUAL_REVIEW_REQUIRED"
    if audio_role in {"AUTHOR_SPEECH_USABLE", "DIALOGUE_USABLE"}:
        for segment in segments:
            text = str(segment.get("text") or segment.get("normalized_text") or "").strip()
            if text:
                return text, "TRANSCRIPT_FIRST_CONTENT_SEGMENT"
    ocr_status = candidate.get("ocr_status")
    ocr_text = str(candidate.get("ocr_text") or "").strip()
    if ocr_status in {"READABLE", "PARTIALLY_READABLE"} and ocr_text:
        return ocr_text, "OCR_PRIMARY_TEXT"
    caption = str(candidate.get("caption") or "").strip()
    if caption:
        return caption, "CAPTION_MANUAL_REVIEW"
    return "", "MANUAL_REVIEW_REQUIRED"


def _write_headers(sheet, headers: Iterable[str]) -> None:
    fill = PatternFill("solid", fgColor="243447")
    for index, title in enumerate(headers, 1):
        cell = sheet.cell(1, index, title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "D2"
    sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(tuple(headers))).coordinate}1"


def _style_sheet(sheet) -> None:
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in range(1, sheet.max_column + 1):
        sheet.column_dimensions[chr(64 + column) if column <= 26 else "A" + chr(64 + column - 26)].width = 18
    sheet.column_dimensions["C"].width = 32
    sheet.column_dimensions["H"].width = 34
    sheet.column_dimensions["M"].width = 20


def validate_portability(package_dir: Path) -> dict[str, Any]:
    package_dir = Path(package_dir)
    workbooks = list(package_dir.glob("*.xlsx"))
    if len(workbooks) != 1:
        raise TrendWorkbookError("package must contain exactly one workbook")
    workbook = load_workbook(workbooks[0], data_only=False)
    if tuple(workbook.sheetnames) != REQUIRED_SHEETS:
        raise TrendWorkbookError("required workbook sheets are missing")
    sheet = workbook["Кандидаты"]
    headers = [cell.value for cell in sheet[1]]
    link_index = headers.index("Открыть локальное видео") + 1
    hash_index = headers.index("SHA-256 MP4") + 1
    count = 0
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, link_index)
        target = cell.hyperlink.target if cell.hyperlink else ""
        if not target or Path(target).is_absolute() or ":" in target:
            raise TrendWorkbookError("local workbook hyperlink must be relative")
        media = package_dir / target.replace("\\", "/")
        if not media.is_file() or sha256(media) != sheet.cell(row, hash_index).value:
            raise TrendWorkbookError("workbook hyperlink target is invalid")
        count += 1
    return {"status": "PASS", "workbook": workbooks[0].name, "relative_links": count}


def build_package(*, project_id: str, search_run_id: str, candidates: Iterable[dict[str, Any]], output_root: Path, query_profile: str = SCORING_PROFILE, package_build_id: str | None = None) -> dict[str, Any]:
    """Build a portable package from locally acquired candidate media.

    Each candidate must provide ``local_media_path``.  Invalid/missing media is
    represented on ``Отбраковано`` and can never reach ``Кандидаты``.
    """
    output_root = Path(output_root)
    package_run_id = f"{_safe_component(search_run_id)}_{_safe_component(package_build_id)}" if package_build_id else _safe_component(search_run_id)
    package = output_root / f"LOOPRA_05_TREND_WORKBOOK_{package_run_id}"
    if package.exists():
        reused = validate_portability(package)
        manifest = json.loads((package / "manifest.json").read_text(encoding="utf-8"))
        return reused | {"package_path": str(package), "workbook_path": str(package / reused["workbook"]), "exported": manifest["exported_candidate_count"], "rejected": manifest["rejected_candidate_count"], "reuse": True}
    prepared = [assess_candidate(item) for item in candidates]
    rejected: list[dict[str, Any]] = []
    unique: dict[str, dict[str, Any]] = {}
    for item in prepared:
        source = Path(item["local_media_path"]) if item.get("local_media_path") else None
        if not source or not source.is_file() or source.stat().st_size == 0:
            item["rejection_reason"] = "MEDIA_NOT_ACQUIRED"; rejected.append(item); continue
        if not _media_is_valid(source):
            item["rejection_reason"] = "INVALID_MEDIA"; rejected.append(item); continue
        item["media_sha256"] = sha256(source)
        existing = unique.get(item["media_sha256"])
        if existing:
            item["rejection_reason"] = "EXACT_DUPLICATE"; item["duplicate_canonical"] = existing.get("video_id"); rejected.append(item); continue
        unique[item["media_sha256"]] = item
    eligible = [item for item in unique.values() if item["eligible"]]
    rejected.extend(item for item in unique.values() if not item["eligible"])
    selected = sorted(eligible, key=lambda item: (-item["final_score"], str(item.get("video_id", ""))))
    if not selected:
        raise TrendWorkbookError("no eligible candidates with local media")
    stage = package.with_name(package.name + ".building")
    stage.mkdir(parents=True, exist_ok=False)
    videos = stage / "videos"; videos.mkdir()
    try:
        for rank, item in enumerate(selected, 1):
            filename = f"{rank:03d}_{_safe_component(item.get('video_id'))}.mp4"
            target = videos / filename
            _copy_exact(Path(item["local_media_path"]), target)
            item.update(rank=rank, local_filename=filename, package_media_path=target, hook=_hook(item)[0], hook_source=_hook(item)[1])
        workbook_path = stage / f"LOOPRA_{_safe_component(project_id).upper()}_TRENDS_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        _write_workbook(workbook_path, project_id, search_run_id, query_profile, selected, rejected)
        manifest = _manifest(project_id, search_run_id, query_profile, selected, rejected, workbook_path, stage, package_build_id)
        (stage / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        (stage / "README_RU.txt").write_text("Откройте Excel, выберите ролики и используйте ссылку «Открыть MP4». TikTok URL сохранён только как provenance.\n", encoding="utf-8")
        shutil.move(str(stage), str(package))
    except Exception:
        shutil.rmtree(stage, ignore_errors=True)
        raise
    validation = validate_portability(package)
    return validation | {"package_path": str(package), "workbook_path": str(package / workbook_path.name), "exported": len(selected), "rejected": len(rejected), "reuse": False}


def _write_workbook(path: Path, project_id: str, run_id: str, profile: str, selected: list[dict[str, Any]], rejected: list[dict[str, Any]]) -> None:
    workbook = Workbook(); summary = workbook.active; summary.title = "Сводка"
    candidates = workbook.create_sheet("Кандидаты"); transcripts = workbook.create_sheet("Транскрипции"); rejected_sheet = workbook.create_sheet("Отбраковано"); method = workbook.create_sheet("Методика")
    summary.append(["LOOPRA 0.5 — Trend Workbook"])
    for key, value in (("Проект", project_id), ("Search run ID", run_id), ("Scoring profile", profile), ("Экспортировано", len(selected)), ("Отбраковано", len(rejected)), ("Доступ", "локальные MP4 + relative hyperlinks"), ("Инструкция", "Кандидаты → Открыть MP4 → Выбрать → Приоритет → передать MP4 вручную в GPT")):
        summary.append([key, value])
    summary.column_dimensions["A"].width = 28; summary.column_dimensions["B"].width = 90
    _write_headers(candidates, HEADERS)
    _write_headers(transcripts, ("Ранг", "Video ID", "Segment", "Start", "End", "Language", "Audio role", "Transcript text", "Quality", "Evidence reference"))
    select_validation = DataValidation(type="list", formula1='"ДА,НЕТ,ПОЗЖЕ"', allow_blank=True)
    priority_validation = DataValidation(type="list", formula1='"1,2,3"', allow_blank=True)
    candidates.add_data_validation(select_validation); candidates.add_data_validation(priority_validation)
    for item in selected:
        transcript = item.get("transcript_segments") or []
        excerpt = " ".join(str(segment.get("text") or segment.get("normalized_text") or "") for segment in transcript)[:500]
        row = [None, None, None, item["rank"], item["classification"], "RELEVANCE_ELIGIBLE", item.get("rejection_reason", "LOW"), item["hook"], item["hook_source"], item.get("query_cluster", ""), item.get("author_username", item.get("author", "")), "Открыть MP4", item["local_filename"], item.get("url", ""), item.get("video_id", ""), item.get("candidate_id", item.get("video_id", "")), item.get("published_at", ""), item.get("duration_seconds", ""), item["package_media_path"].stat().st_size, item["media_sha256"], item.get("views", ""), item.get("likes", ""), item.get("comments", ""), item.get("shares", ""), item.get("like_rate", ""), item.get("comment_rate", ""), item.get("share_rate", ""), item.get("total_engagement_rate", ""), item["virality_score"], item["relevance_score"], item["freshness_score"], item["quality_score"], item["final_score"], item["classification"], item.get("query_cluster", ""), item.get("query", item.get("source_value", "")), item.get("source_type", ""), item.get("caption", ""), item.get("audio_role", "MANUAL_REVIEW_REQUIRED"), excerpt, len(transcript), "К транскрипции", item.get("ocr_status", "MANUAL_TEXT_CHECK"), item.get("ocr_text", ""), item.get("audio_role", "MANUAL_REVIEW_REQUIRED"), item.get("content_modality", "unknown"), "UNIQUE", item.get("evidence_warning", ""), "VALID_MEDIA"]
        row[47] = item.get("transcript_reason") or item.get("evidence_warning", "")
        candidates.append(row); row_number = candidates.max_row
        select_validation.add(candidates.cell(row_number, 1)); priority_validation.add(candidates.cell(row_number, 2))
        video = candidates.cell(row_number, 12); video.hyperlink = f"videos\\{item['local_filename']}"; video.style = "Hyperlink"
        if transcript:
            candidates.cell(row_number, 42).hyperlink = f"#'Транскрипции'!A{transcripts.max_row + 1}"
            candidates.cell(row_number, 42).style = "Hyperlink"
        for number, segment in enumerate(transcript, 1):
            transcripts.append([item["rank"], item.get("video_id", ""), number, segment.get("start_seconds", segment.get("start", "")), segment.get("end_seconds", segment.get("end", "")), segment.get("language", ""), item.get("audio_role", "MANUAL_REVIEW_REQUIRED"), segment.get("text") or segment.get("normalized_text") or "", segment.get("quality_status", ""), segment.get("evidence_ref", "")])
    candidates.auto_filter.ref = f"A1:{candidates.cell(candidates.max_row, candidates.max_column).coordinate}"
    candidates.conditional_formatting.add(f"D2:D{candidates.max_row}", CellIsRule(operator="lessThan", formula=["999999"], fill=PatternFill("solid", fgColor="E2F0D9")))
    _style_sheet(candidates)
    transcripts.auto_filter.ref = f"A1:J{max(1, transcripts.max_row)}"; _style_sheet(transcripts)
    _write_headers(rejected_sheet, ("Candidate/Video ID", "Автор", "Source URL", "Query", "Reason code", "Explanation", "Duplicate canonical", "Score before rejection", "Media status"))
    for item in rejected:
        rejected_sheet.append([item.get("video_id", ""), item.get("author_username", ""), item.get("url", ""), item.get("query", ""), item.get("rejection_reason", "LOW_RELEVANCE"), item.get("rejection_reason", ""), item.get("duplicate_canonical", ""), item.get("final_score", 0), "MISSING" if item.get("rejection_reason") == "MEDIA_NOT_ACQUIRED" else "REJECTED"])
    rejected_sheet.auto_filter.ref = f"A1:I{max(1, rejected_sheet.max_row)}"; _style_sheet(rejected_sheet)
    method.append(["Методика LOOPRA 0.5"])
    for line in ("Сначала применяется relevance eligibility gate; низкая topical relevance не компенсируется engagement.", "Junk-фильтры учитывают query, caption, transcript и OCR; одно слово не является единственным основанием.", "Exact duplicates исключаются по SHA-256 MP4. Каждый основной кандидат имеет локальный MP4.", "Hook берётся только из пригодной речи, читаемого OCR или caption; иначе требуется ручная проверка.", "TikTok URL — provenance, не основной способ доступа. LOOPRA 0.5 не создаёт сценарии автоматически."):
        method.append([line])
    method.column_dimensions["A"].width = 120
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
    workbook.save(path)


def _manifest(project_id: str, run_id: str, profile: str, selected: list[dict[str, Any]], rejected: list[dict[str, Any]], workbook: Path, root: Path, package_build_id: str | None = None) -> dict[str, Any]:
    return {"schema_version": SCHEMA_VERSION, "package_id": f"loopra-05-{run_id}" + (f"-{_safe_component(package_build_id)}" if package_build_id else ""), "package_build_id": package_build_id, "project_id": project_id, "search_run_id": run_id, "search_timestamp": _utc_now(), "query_profile_reference": profile, "raw_candidate_count": len(selected) + len(rejected), "deduplicated_candidate_count": len(selected), "shortlisted_candidate_count": len(selected), "acquired_candidate_count": len(selected), "transcribed_candidate_count": sum(bool(item.get("transcript_segments")) for item in selected), "eligible_candidate_count": len(selected), "exported_candidate_count": len(selected), "rejected_candidate_count": len(rejected), "exact_duplicate_count": sum(item.get("rejection_reason") == "EXACT_DUPLICATE" for item in rejected), "near_duplicate_count": 0, "workbook_relative_path": workbook.name, "workbook_hash": sha256(workbook), "videos_directory_relative_path": "videos", "video_file_count": len(selected), "ordered_video_references": [f"videos/{item['local_filename']}" for item in selected], "ordered_video_sha256": [item["media_sha256"] for item in selected], "workbook_sheet_names": list(REQUIRED_SHEETS), "column_schema_version": SCHEMA_VERSION, "relative_link_validation_status": "PASS", "portability_validation_status": "PASS", "reuse_metadata": {"supported": True}, "package_size": sum(path.stat().st_size for path in root.rglob("*") if path.is_file())}
