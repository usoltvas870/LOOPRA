#!/usr/bin/env python3
"""One-shot, isolated acceptance harness for the Trend Radar Data Trust slice."""

from __future__ import annotations

import argparse
import json
import os
import signal
import subprocess
import sys
import time
from datetime import datetime, timezone
from pathlib import Path


RADAR_ROOT = Path(__file__).resolve().parent
DEFAULT_TIMEOUT_SECONDS = 300
EXPECTED_SHEETS = {'Videos', 'Source Coverage', 'Provenance Matches', 'Run Summary'}
FORBIDDEN_EVIDENCE_MARKERS = (
    'authorization:',
    'sessionid=',
    '"sessionid"',
    'mstoken=',
    '"mstoken"',
    'sk-your-key',
)


def _utc_stamp() -> str:
    return datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')


def _write_json(path: Path, value: object) -> None:
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding='utf-8')


def _related_processes(evidence_dir: Path) -> list[dict]:
    if os.name != 'nt':
        return []
    escaped = str(evidence_dir).replace("'", "''")
    script = (
        "$items=Get-CimInstance Win32_Process | "
        f"Where-Object {{ $_.ProcessId -ne $PID -and $_.CommandLine -like '*{escaped}*' }} | "
        "Select-Object ProcessId,ParentProcessId,Name,ExecutablePath,CreationDate,CommandLine; "
        "if($items){$items|ConvertTo-Json -Compress}"
    )
    result = subprocess.run(
        ['powershell', '-NoProfile', '-Command', script],
        capture_output=True,
        text=True,
        encoding='utf-8',
        errors='replace',
        timeout=15,
        check=False,
    )
    if not result.stdout.strip():
        return []
    parsed = json.loads(result.stdout)
    return parsed if isinstance(parsed, list) else [parsed]


def _terminate_spawned_tree(process: subprocess.Popen, grace_seconds: int = 10) -> str:
    if process.poll() is not None:
        return 'already_exited'
    if os.name == 'nt':
        try:
            process.send_signal(signal.CTRL_BREAK_EVENT)
            process.wait(timeout=grace_seconds)
            return 'ctrl_break'
        except (OSError, subprocess.TimeoutExpired):
            subprocess.run(
                ['taskkill', '/PID', str(process.pid), '/T', '/F'],
                capture_output=True,
                text=True,
                check=False,
            )
            try:
                process.wait(timeout=10)
            except subprocess.TimeoutExpired:
                pass
            return 'forced_process_tree'
    process.terminate()
    try:
        process.wait(timeout=grace_seconds)
        return 'terminate'
    except subprocess.TimeoutExpired:
        process.kill()
        process.wait(timeout=10)
        return 'forced_process_tree'


def _scan_evidence(evidence_dir: Path, secret_values: list[str]) -> list[str]:
    findings = []
    for path in evidence_dir.rglob('*'):
        if not path.is_file() or path.suffix.lower() in {'.xlsx', '.db'}:
            continue
        text = path.read_text(encoding='utf-8', errors='replace').lower()
        for marker in FORBIDDEN_EVIDENCE_MARKERS:
            if marker in text:
                findings.append(f'{path.name}: forbidden marker {marker}')
        for value in secret_values:
            if len(value) >= 8 and value.lower() in text:
                findings.append(f'{path.name}: secret environment value')
    return findings


def validate_success_artifacts(evidence_dir: Path, child_result: dict) -> tuple[list[str], dict]:
    failures = []
    stdout = (evidence_dir / 'stdout.log').read_text(encoding='utf-8', errors='replace')
    radar_results = [line.strip() for line in stdout.splitlines() if line.strip().startswith('RADAR_RESULT=')]
    if radar_results != ['RADAR_RESULT=success']:
        failures.append(f'expected one success RADAR_RESULT, got {radar_results}')
    if child_result.get('exit_code') != 0:
        failures.append(f'expected exit code 0, got {child_result.get("exit_code")}')
    if child_result.get('timeout'):
        failures.append('child timed out')
    if child_result.get('termination_reason') != 'normal_exit':
        failures.append(f'unexpected termination: {child_result.get("termination_reason")}')

    journals = sorted((evidence_dir / 'journal').glob('run_*.json'))
    reports = sorted((evidence_dir / 'reports').glob('report_*.md'))
    workbooks = sorted((evidence_dir / 'reports').glob('report_*.xlsx'))
    if len(journals) != 1:
        failures.append(f'expected one journal, got {len(journals)}')
        return failures, {}
    journal = json.loads(journals[0].read_text(encoding='utf-8'))
    attempts = journal.get('source_attempts', [])
    provenance = journal.get('provenance', [])
    if journal.get('sources_processed') != 3 or len(attempts) != 3:
        failures.append('planned/attempted source count is not 3')
    required_attempt_fields = (
        'started_at', 'completed_at', 'duration_ms', 'requested_limit',
        'raw_items_received', 'parsed_items', 'collection_method', 'authentication_state',
    )
    for attempt in attempts:
        missing = [field for field in required_attempt_fields if attempt.get(field) is None]
        if missing:
            failures.append(f'source {attempt.get("ordinal")} missing {missing}')
        if attempt.get('requested_limit', 99) > 3:
            failures.append(f'source {attempt.get("ordinal")} exceeded requested limit')
    for item in provenance:
        for field in ('video_id', 'canonical_url', 'matched_sources', 'new_to_database'):
            if item.get(field) is None:
                failures.append(f'provenance missing {field}')
    run_videos = journal.get('run_videos', [])
    if len(run_videos) != journal.get('unique_discoveries'):
        failures.append('run video evidence does not match unique discoveries')
    for item in run_videos:
        if not item.get('freshness_bucket'):
            failures.append(f'video {item.get("video_id")} missing freshness bucket')
        if item.get('published_at') and item.get('age_hours') is None:
            failures.append(f'video {item.get("video_id")} has publication date without age')

    if len(reports) != 1 or not reports[0].stat().st_size:
        failures.append(f'expected one non-empty Markdown report, got {len(reports)}')
    else:
        markdown = reports[0].read_text(encoding='utf-8')
        for heading in ('Run Summary', 'Source Coverage', 'Freshness Summary', 'Data Quality Warnings'):
            if heading not in markdown:
                failures.append(f'Markdown missing {heading}')
    if len(workbooks) != 1:
        failures.append(f'expected one XLSX report, got {len(workbooks)}')
    else:
        from openpyxl import load_workbook
        workbook = load_workbook(workbooks[0], read_only=True)
        if set(workbook.sheetnames) != EXPECTED_SHEETS:
            failures.append(f'unexpected XLSX sheets: {workbook.sheetnames}')
        workbook.close()

    _write_json(evidence_dir / 'provenance.json', provenance)
    summary = {
        'journal': str(journals[0]),
        'markdown': str(reports[0]) if reports else None,
        'xlsx': str(workbooks[0]) if workbooks else None,
        'attempts': attempts,
        'provenance_count': len(provenance),
        'freshness_buckets': journal.get('freshness_summary', {}),
    }
    return failures, summary


def run_acceptance(timeout_seconds: int) -> tuple[Path, int]:
    evidence_dir = RADAR_ROOT / 'data' / 'evidence' / f'data_trust_acceptance_{_utc_stamp()}'
    config_dir = evidence_dir / 'config'
    reports_dir = evidence_dir / 'reports'
    journal_dir = evidence_dir / 'journal'
    temp_dir = evidence_dir / 'temp'
    for directory in (config_dir, reports_dir, journal_dir, temp_dir):
        directory.mkdir(parents=True, exist_ok=True)

    (config_dir / 'competitors.txt').write_text('', encoding='utf-8')
    (config_dir / 'hashtags.txt').write_text('матрицасудьбы\nnumerology\n', encoding='utf-8')
    (config_dir / 'keywords.txt').write_text('предназначение по дате рождения\n', encoding='utf-8')
    (config_dir / 'rotational.txt').write_text('', encoding='utf-8')

    sanitized_config = {
        'sources': [
            {'type': 'hashtag', 'value': 'матрицасудьбы'},
            {'type': 'hashtag', 'value': 'numerology'},
            {'type': 'keyword', 'value': 'предназначение по дате рождения'},
        ],
        'MAX_RESULTS_PER_SOURCE': 3,
        'HEADLESS': False,
        'LOG_LEVEL': 'INFO',
        'ENABLE_AI_ANALYSIS': False,
        'ENABLE_TELEGRAM': False,
        'timeout_seconds': timeout_seconds,
    }
    _write_json(evidence_dir / 'sanitized_config.json', sanitized_config)
    (evidence_dir / 'commands.md').write_text(
        '# Acceptance command\n\n'
        f'`python trend-radar/run_acceptance.py --timeout {timeout_seconds}`\n\n'
        'The harness starts `run_radar.py` with isolated source, report, journal, stdout, and stderr paths.\n',
        encoding='utf-8',
    )

    child_env = os.environ.copy()
    child_env.update({
        'SOURCE_CONFIG_DIR': str(config_dir),
        'RADAR_REPORTS_DIR': str(reports_dir),
        'RADAR_EVIDENCE_DIR': str(journal_dir),
        'COOKIE_PATH': str(RADAR_ROOT / 'data' / 'tiktok_cookies.json'),
        'MAX_RESULTS_PER_SOURCE': '3',
        'HEADLESS': 'false',
        'LOG_LEVEL': 'INFO',
        'ENABLE_AI_ANALYSIS': 'false',
        'ENABLE_TELEGRAM': 'false',
        'EXPORT_XLSX': 'true',
        'TMP': str(temp_dir),
        'TEMP': str(temp_dir),
    })
    secret_values = [
        value for key, value in child_env.items()
        if any(token in key.upper() for token in ('TOKEN', 'SECRET', 'API_KEY', 'PASSWORD'))
        and value
    ]

    stdout_path = evidence_dir / 'stdout.log'
    stderr_path = evidence_dir / 'stderr.log'
    started = time.monotonic()
    creation_flags = subprocess.CREATE_NEW_PROCESS_GROUP if os.name == 'nt' else 0
    with stdout_path.open('w', encoding='utf-8') as stdout_file, stderr_path.open('w', encoding='utf-8') as stderr_file:
        process = subprocess.Popen(
            [sys.executable, str(RADAR_ROOT / 'run_radar.py')],
            cwd=RADAR_ROOT,
            env=child_env,
            stdout=stdout_file,
            stderr=stderr_file,
            text=True,
            creationflags=creation_flags,
        )
        timed_out = False
        termination_reason = 'normal_exit'
        try:
            exit_code = process.wait(timeout=timeout_seconds)
        except subprocess.TimeoutExpired:
            timed_out = True
            termination_reason = _terminate_spawned_tree(process)
            exit_code = process.poll()

    duration_seconds = round(time.monotonic() - started, 3)
    lingering = _related_processes(evidence_dir)
    child_result = {
        'pid': process.pid,
        'exit_code': exit_code,
        'timeout': timed_out,
        'timeout_seconds': timeout_seconds,
        'termination_reason': termination_reason,
        'duration_seconds': duration_seconds,
        'lingering_related_processes': lingering,
    }
    _write_json(evidence_dir / 'process_result.json', child_result)

    failures, summary = validate_success_artifacts(evidence_dir, child_result)
    if lingering:
        failures.append(f'{len(lingering)} acceptance-related process(es) remain')
    security_findings = _scan_evidence(evidence_dir, secret_values)
    failures.extend(security_findings)
    assertions = {'passed': not failures, 'failures': failures, 'security_findings': security_findings, 'summary': summary}
    _write_json(evidence_dir / 'assertions.json', assertions)

    report_lines = [
        '# Trend Radar Data Trust Acceptance',
        '',
        f'**Status:** {"PASS" if not failures else "PARTIAL"}',
        f'**Timeout:** {timeout_seconds} seconds',
        f'**Duration:** {duration_seconds} seconds',
        f'**Exit code:** {exit_code}',
        f'**Timeout flag:** {timed_out}',
        f'**Termination reason:** {termination_reason}',
        '',
        '## Assertions',
        '',
    ]
    report_lines += ['- All assertions passed.'] if not failures else [f'- {failure}' for failure in failures]
    report_lines += ['', '## Source attempts', '']
    for attempt in summary.get('attempts', []):
        report_lines.append(
            f'- {attempt.get("source_type")}/{attempt.get("source_value")}: '
            f'{attempt.get("status")}, raw={attempt.get("raw_items_received")}, '
            f'parsed={attempt.get("parsed_items")}, duration_ms={attempt.get("duration_ms")}'
        )
    report_lines += [
        '',
        '## Data summary',
        '',
        f'- Unique discoveries: {summary.get("provenance_count", 0)}',
        f'- Freshness buckets: {json.dumps(summary.get("freshness_buckets", {}), ensure_ascii=False)}',
        f'- Cross-source overlap is allowed to be zero; provenance structure is still asserted.',
        '',
        '## Artifacts',
        '',
        f'- Markdown: {summary.get("markdown")}',
        f'- XLSX: {summary.get("xlsx")}',
        f'- JSON journal: {summary.get("journal")}',
        f'- Provenance: {evidence_dir / "provenance.json"}',
    ]
    report_lines += [
        '',
        '## Cleanup',
        '',
        f'- Remaining acceptance-related processes: {len(lingering)}',
        '- Ordinary user Chrome processes were not targeted.',
        '',
        '## Security',
        '',
        f'- Secret scan findings: {len(security_findings)}',
        '- Cookie storage was referenced in place and was not copied into evidence.',
    ]
    (evidence_dir / 'acceptance_report.md').write_text('\n'.join(report_lines) + '\n', encoding='utf-8')
    return evidence_dir, 0 if not failures else 1


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument('--timeout', type=int, default=DEFAULT_TIMEOUT_SECONDS)
    args = parser.parse_args()
    if args.timeout < 60:
        parser.error('timeout must be at least 60 seconds')
    evidence_dir, exit_code = run_acceptance(args.timeout)
    print(f'ACCEPTANCE_EVIDENCE={evidence_dir}')
    print(f'ACCEPTANCE_RESULT={"pass" if exit_code == 0 else "partial"}')
    return exit_code


if __name__ == '__main__':
    raise SystemExit(main())
