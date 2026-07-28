from __future__ import annotations

import json
import shutil
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))
import trend_workbook


@pytest.fixture(autouse=True)
def valid_media(monkeypatch):
    monkeypatch.setattr(trend_workbook, "_media_is_valid", lambda path: True)


def candidate(video_id: str, media: Path | None, **extra):
    return {
        "video_id": video_id, "candidate_id": f"candidate-{video_id}", "local_media_path": str(media) if media else None,
        "topical_relevance": 0.9, "audience_relevance": 0.8, "transferable_potential": 0.7,
        "virality_score": 0.6, "freshness_score": 0.5, "evidence_quality": 0.8,
        "caption": "Это осмысленная тема про личные границы", "query": "личные границы", "query_cluster": "личные границы",
        "url": f"https://www.tiktok.com/@example/video/{video_id}", "audio_role": "AUTHOR_SPEECH_USABLE",
        "transcript_segments": [{"start_seconds": 0, "end_seconds": 3, "text": "Я долго поддерживала всех и забывала о себе."}],
        "ocr_status": "GARBLED", "ocr_text": "???", **extra,
    }


def test_package_has_portable_relative_links_and_operator_workbook(tmp_path: Path):
    source = tmp_path / "source.mp4"; source.write_bytes(b"a valid test byte stream")
    result = trend_workbook.build_package(project_id="nura", search_run_id="run-01", candidates=[candidate("123", source)], output_root=tmp_path / "output")
    assert result["exported"] == 1 and result["relative_links"] == 1
    package = Path(result["package_path"]); workbook = load_workbook(result["workbook_path"])
    assert tuple(workbook.sheetnames) == trend_workbook.REQUIRED_SHEETS
    sheet = workbook["Кандидаты"]
    assert [cell.value for cell in sheet[1]][:3] == list(trend_workbook.OPERATOR_COLUMNS)
    assert sheet.freeze_panes == "D2" and sheet.auto_filter.ref
    target = sheet.cell(2, 12).hyperlink.target
    assert target == "videos\\001_123.mp4" and not Path(target).is_absolute()
    assert (package / target.replace("\\", "/")).is_file()
    assert sheet.cell(2, 8).value == "Я долго поддерживала всех и забывала о себе."
    assert sheet.cell(2, 9).value == "TRANSCRIPT_FIRST_CONTENT_SEGMENT"
    assert sheet.cell(2, 42).hyperlink.target == "#'Транскрипции'!A2"
    transcript_sheet = workbook["Транскрипции"]
    assert transcript_sheet.max_row == 2
    assert transcript_sheet.cell(2, 8).value == "Я долго поддерживала всех и забывала о себе."


def test_rejections_keep_missing_media_duplicate_and_low_relevance_out_of_main_sheet(tmp_path: Path):
    media = tmp_path / "source.mp4"; media.write_bytes(b"same media")
    other = tmp_path / "other.mp4"; other.write_bytes(b"other media")
    result = trend_workbook.build_package(project_id="nura", search_run_id="run-02", candidates=[candidate("a", media), candidate("b", media), candidate("c", None), candidate("d", other, topical_relevance=0.1)], output_root=tmp_path / "output")
    workbook = load_workbook(result["workbook_path"])
    assert workbook["Кандидаты"].max_row == 2
    reasons = {workbook["Отбраковано"].cell(row, 5).value for row in range(2, workbook["Отбраковано"].max_row + 1)}
    assert {"EXACT_DUPLICATE", "MEDIA_NOT_ACQUIRED", "LOW_RELEVANCE"} <= reasons


def test_low_relevance_cannot_be_rescued_by_virality():
    item = trend_workbook.assess_candidate({"topical_relevance": 0.1, "virality_score": 100.0})
    assert item["eligible"] is False and item["final_score"] == 0


def test_portability_detects_absolute_link(tmp_path: Path):
    source = tmp_path / "source.mp4"; source.write_bytes(b"media")
    result = trend_workbook.build_package(project_id="nura", search_run_id="run-03", candidates=[candidate("123", source)], output_root=tmp_path / "output")
    workbook = load_workbook(result["workbook_path"]); workbook["Кандидаты"].cell(2, 12).hyperlink = "C:\\absolute.mp4"; workbook.save(result["workbook_path"])
    with pytest.raises(trend_workbook.TrendWorkbookError, match="relative"):
        trend_workbook.validate_portability(Path(result["package_path"]))


def test_package_reuse_is_deterministic(tmp_path: Path):
    source = tmp_path / "source.mp4"; source.write_bytes(b"media")
    first = trend_workbook.build_package(project_id="nura", search_run_id="run-04", candidates=[candidate("123", source)], output_root=tmp_path / "output")
    second = trend_workbook.build_package(project_id="nura", search_run_id="run-04", candidates=[], output_root=tmp_path / "output")
    assert second["reuse"] is True and second["workbook_path"] == first["workbook_path"]
    manifest = json.loads((Path(first["package_path"]) / "manifest.json").read_text(encoding="utf-8"))
    assert manifest["schema_version"] == trend_workbook.SCHEMA_VERSION


def test_public_workflow_uses_twenty_v2_per_item_calls_and_never_provider(tmp_path: Path, monkeypatch):
    media = []
    for rank in range(1, 21):
        path = tmp_path / f"{rank}.mp4"; path.write_bytes(f"media-{rank}".encode()); media.append(path)
    calls = {"acquire": 0, "ocr": 0, "transcribe": 0, "close": 0}
    pool = {"search_run_id": "public-run", "public_access_status": "PUBLIC_ACCESS_SUFFICIENT", "candidates": [candidate(str(rank), media[rank - 1], query="личные границы") for rank in range(1, 21)]}
    entries = [{"candidate_id": f"c{rank}", "video_id": str(rank), "original_rank": rank} for rank in range(1, 21)]
    def acquire(entry):
        calls["acquire"] += 1
        source = media[entry["original_rank"] - 1]
        return {"status": "COMPLETED", "ffprobe_status": "VALID", "source_media_reference": source.relative_to(tmp_path).as_posix()}
    deps = {"collect": lambda: pool, "select": lambda _: entries, "acquire": acquire, "ocr": lambda _: calls.__setitem__("ocr", calls["ocr"] + 1) or {}, "transcribe": lambda _: calls.__setitem__("transcribe", calls["transcribe"] + 1) or {}, "close": lambda: calls.__setitem__("close", calls["close"] + 1)}
    result = trend_workbook.run_public_first_workbook(project_id="nura", runtime_root=tmp_path, output_root=tmp_path / "output", production_dependencies=deps)
    assert result["status"] == "READY_FOR_OWNER_WORKBOOK_REVIEW"
    assert calls == {"acquire": 20, "ocr": 20, "transcribe": 20, "close": 1}
    assert result["provider_calls"] == result["script_calls"] == 0


def test_public_workflow_continues_after_one_media_failure(tmp_path: Path):
    media = []
    for rank in range(1, 21):
        path = tmp_path / f"{rank}.mp4"; path.write_bytes(f"media-{rank}".encode()); media.append(path)
    pool = {"search_run_id": "partial-run", "public_access_status": "PUBLIC_ACCESS_SUFFICIENT", "candidates": [candidate(str(rank), media[rank - 1]) for rank in range(1, 21)]}
    entries = [{"candidate_id": f"c{rank}", "video_id": str(rank), "original_rank": rank} for rank in range(1, 21)]
    acquired = []
    def acquire(entry):
        acquired.append(entry["original_rank"])
        if entry["original_rank"] == 6: return {"status": "FAILED", "ffprobe_status": "INVALID"}
        return {"status": "COMPLETED", "ffprobe_status": "VALID", "source_media_reference": media[entry["original_rank"] - 1].relative_to(tmp_path).as_posix()}
    deps = {"collect": lambda: pool, "select": lambda _: entries, "acquire": acquire, "ocr": lambda _: {}, "transcribe": lambda _: {}, "close": lambda: None}
    result = trend_workbook.run_public_first_workbook(project_id="nura", runtime_root=tmp_path, output_root=tmp_path / "output", production_dependencies=deps)
    assert acquired == list(range(1, 21))
    assert result["status"] == "PARTIAL_INSUFFICIENT_VALID_MEDIA"


def test_backfill_target_counts_unique_valid_media_not_attempts(tmp_path: Path):
    items = [candidate(str(rank), None, query="личные границы") for rank in range(1, 31)]
    paths = {}
    for rank in range(1, 31):
        path = tmp_path / f"{rank}.mp4"; path.write_bytes(f"media-{rank}".encode()); paths[rank] = path
    # Existing positions 10/19 failed and 18 duplicates position 13: 17 unique.
    def lookup(item):
        rank = int(item["video_id"])
        if rank > 20: return None
        if rank in {10, 19}: return {"status": "FAILED", "reason": "MEDIA_NOT_ACQUIRED"}
        path = paths[13] if rank == 18 else paths[rank]
        return {"status": "REUSED", "local_media_path": str(path), "media_sha256": trend_workbook.sha256(path)}
    acquired = []
    def acquire(item, position):
        acquired.append(position); path = paths[position]
        return {"status": "COMPLETED", "local_media_path": str(path), "media_sha256": trend_workbook.sha256(path)}
    result = trend_workbook.backfill_candidate_media(candidates=items, lookup_media=lookup, acquire_one=acquire, target_count=20)
    assert result["starting_unique_media"] == 17
    assert result["valid_unique_media"] == 20
    assert result["additional_attempts"] == 3 and acquired == [21, 22, 23]
    assert result["failures"] == 2 and result["duplicates"] == 1
    assert result["processed_positions"][-1] == 23


def test_backfill_stops_safely_when_shortlist_is_exhausted(tmp_path: Path):
    items = [candidate(str(rank), None) for rank in range(1, 4)]
    result = trend_workbook.backfill_candidate_media(
        candidates=items, lookup_media=lambda _: None,
        acquire_one=lambda item, position: {"status": "FAILED", "reason": "MEDIA_NOT_ACQUIRED"},
        target_count=2, maximum_attempts=10, maximum_shortlist_size=3,
    )
    assert result["valid_unique_media"] == 0 and result["additional_attempts"] == 3
    assert result["shortlist_exhausted"] is True


def test_backfill_reuses_existing_media_without_redownload(tmp_path: Path):
    items = [candidate(str(rank), None) for rank in range(1, 4)]
    paths = []
    for rank in range(1, 4):
        path = tmp_path / f"{rank}.mp4"; path.write_bytes(str(rank).encode()); paths.append(path)
    def lookup(item):
        path = paths[int(item["video_id"]) - 1]
        return {"status": "REUSED", "local_media_path": str(path), "media_sha256": trend_workbook.sha256(path)}
    result = trend_workbook.backfill_candidate_media(
        candidates=items, lookup_media=lookup,
        acquire_one=lambda *_: pytest.fail("existing media must not be downloaded"), target_count=3,
    )
    assert result["valid_unique_media"] == 3 and result["additional_attempts"] == 0


@pytest.mark.parametrize(("payload", "caption", "expected"), [
    ({"status": "COMPLETED_NO_AUDIO", "segments": []}, "", "NO_AUDIO"),
    ({"status": "COMPLETED_NO_SPEECH", "segments": []}, "", "NONSPEECH"),
    ({"status": "FAILED", "errors": ["decoder failed"]}, "", "TRANSCRIPTION_FAILED"),
    ({"status": "COMPLETED", "language": "ru", "segments": [{"normalized_text": "ла ла ла песня", "start_seconds": 0, "end_seconds": 1}]}, "#music", "BACKGROUND_MUSIC_ONLY"),
    ({"status": "COMPLETED", "language": "ru", "segments": [{"normalized_text": "Это осмысленная человеческая речь", "start_seconds": 0, "end_seconds": 2}]}, "", "AUTHOR_SPEECH_USABLE"),
])
def test_transcription_status_is_truthful_and_typed(payload, caption, expected):
    result = trend_workbook._classify_transcription(payload, "evidence/result.json", caption)
    assert result["audio_role"] == expected and result["transcript_reason"]
    if expected == "BACKGROUND_MUSIC_ONLY":
        assert trend_workbook._hook({**result, "caption": "", "ocr_status": "EMPTY"})[1] == "MANUAL_REVIEW_REQUIRED"


def test_canonical_transcription_is_called_and_segments_are_preserved(tmp_path: Path):
    media = tmp_path / "canonical/acquisition/run-1/video-1/source.mp4"
    media.parent.mkdir(parents=True); media.write_bytes(b"media")
    digest = trend_workbook.sha256(media)
    (media.parent / "acquisition_record.json").write_text(json.dumps({"local_media_path": "video-1/source.mp4", "media_sha256": digest}), encoding="utf-8")
    candidate_item = candidate("video-1", media)
    def inspection(source, output, video_id, url):
        output.mkdir(parents=True)
        payload = {"media_sha256": digest, "status": "COMPLETED", "media_facts": {"audio_present": True, "audio_codec": "aac", "duration_seconds": 3, "sample_rate": "44100", "channels": 1}}
        (output / "inspection.json").write_text(json.dumps(payload), encoding="utf-8")
        return payload
    class Engine:
        calls = 0
        def availability(self): return {"available": True, "engine_id": "fake", "engine_version": "1", "model_id": "fake", "model_revision": "1"}
        def transcribe(self, media_path, language, options):
            self.calls += 1
            return {"language": "ru", "language_probability": .99, "segments": [{"start_seconds": 0, "end_seconds": 2, "raw_text": " Полная человеческая речь сохранена ", "avg_logprob": -.1, "no_speech_prob": .01, "words": []}]}
    engine = Engine()
    evidence = trend_workbook._build_transcription_evidence(
        runtime_root=tmp_path, build_root=tmp_path / "build", search_run_id="run-1",
        candidates=[candidate_item], inspection_callable=inspection, engine=engine,
    )
    assert engine.calls == 1 and evidence["transcription_calls"] == 1
    result = evidence["candidates"][0]
    assert result["audio_role"] == "AUTHOR_SPEECH_USABLE"
    assert result["transcript_segments"][0]["text"] == "Полная человеческая речь сохранена"


def test_versioned_package_does_not_mutate_old_partial(tmp_path: Path):
    source = tmp_path / "source.mp4"; source.write_bytes(b"media")
    old = trend_workbook.build_package(project_id="nura", search_run_id="run", candidates=[candidate("1", source)], output_root=tmp_path / "output")
    old_hash = trend_workbook.sha256(Path(old["workbook_path"]))
    new = trend_workbook.build_package(project_id="nura", search_run_id="run", package_build_id="recovery", candidates=[candidate("1", source)], output_root=tmp_path / "output")
    assert Path(old["package_path"]) != Path(new["package_path"])
    assert trend_workbook.sha256(Path(old["workbook_path"])) == old_hash


def test_twenty_candidate_package_remains_portable_after_relocation(tmp_path: Path):
    items = []
    for rank in range(1, 21):
        source = tmp_path / "sources" / f"{rank}.mp4"
        source.parent.mkdir(exist_ok=True)
        source.write_bytes(f"media-{rank}".encode())
        items.append(candidate(str(rank), source))
    result = trend_workbook.build_package(
        project_id="nura", search_run_id="run-20", package_build_id="recovery",
        candidates=items, output_root=tmp_path / "output",
    )
    relocated = tmp_path / "relocated" / Path(result["package_path"]).name
    relocated.parent.mkdir()
    shutil.copytree(result["package_path"], relocated)
    validation = trend_workbook.validate_portability(relocated)
    workbook = load_workbook(next(relocated.glob("*.xlsx")))
    assert validation == {"status": "PASS", "workbook": next(relocated.glob("*.xlsx")).name, "relative_links": 20}
    assert workbook["Кандидаты"].max_row == 21
    assert len(list((relocated / "videos").glob("*.mp4"))) == 20


def test_resume_reuses_versioned_package_without_acquisition_or_transcription(tmp_path: Path):
    runtime = tmp_path / "runtime"
    collection = {"search_run_id": "run-reuse", "candidates": []}
    (runtime / "canonical").mkdir(parents=True)
    (runtime / "canonical" / "collection.json").write_text(json.dumps(collection), encoding="utf-8")
    source = tmp_path / "source.mp4"; source.write_bytes(b"media")
    package = trend_workbook.build_package(
        project_id="nura", search_run_id="run-reuse", package_build_id="stable",
        candidates=[candidate("1", source)], output_root=tmp_path / "output",
    )
    manifest_path = Path(package["package_path"]) / "manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["transcription"] = {"status_counts": {"AUTHOR_SPEECH_USABLE": 1}}
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    result = trend_workbook.resume_public_first_workbook(
        project_id="nura", runtime_root=runtime, output_root=tmp_path / "output",
        target_count=20, build_id="stable",
        acquire_one=lambda *_: pytest.fail("reuse must not acquire"),
        transcription_engine=object(),
    )
    assert result["reuse"] is True
    assert result["search_calls"] == result["browser_calls"] == result["downloads"] == result["transcription_calls"] == 0


def test_resume_reports_canonical_transcription_execution_count(tmp_path: Path, monkeypatch):
    runtime = tmp_path / "runtime"
    (runtime / "canonical").mkdir(parents=True)
    items = []
    for rank in range(1, 21):
        source = tmp_path / "sources" / f"{rank}.mp4"
        source.parent.mkdir(exist_ok=True)
        source.write_bytes(f"media-{rank}".encode())
        items.append(candidate(str(rank), source))
    collection = {"search_run_id": "run-count", "candidates": items}
    (runtime / "canonical" / "collection.json").write_text(json.dumps(collection), encoding="utf-8")
    monkeypatch.setattr(trend_workbook, "backfill_candidate_media", lambda **_: {
        "candidates": items, "starting_unique_media": 20, "additional_attempts": 0,
        "failures": 0, "duplicates": 0, "valid_unique_media": 20,
        "processed_positions": list(range(1, 21)), "shortlist_exhausted": False,
    })
    monkeypatch.setattr(trend_workbook, "_build_transcription_evidence", lambda **_: {
        "candidates": items, "transcription_calls": 20, "transcription_reused": 0,
        "status_counts": {"AUTHOR_SPEECH_USABLE": 20}, "manifest_path": "selection.json",
    })
    result = trend_workbook.resume_public_first_workbook(
        project_id="nura", runtime_root=runtime, output_root=tmp_path / "output",
        target_count=20, build_id="count", acquire_one=lambda *_: None,
    )
    assert result["status"] == "READY_FOR_OWNER_WORKBOOK_REVIEW"
    assert result["transcription_calls"] == 20


def test_workbook_preserves_large_video_ids_and_bounds_display_excerpts(tmp_path: Path):
    source = tmp_path / "source.mp4"; source.write_bytes(b"media")
    video_id = 7666898189458264686
    item = candidate(str(video_id), source, caption="длинный caption " * 100)
    item["video_id"] = video_id
    item["candidate_id"] = video_id
    item["transcript_segments"] = [{"start_seconds": 0, "end_seconds": 2, "text": "длинная речь " * 100}]
    result = trend_workbook.build_package(
        project_id="nura", search_run_id="large-id", candidates=[item], output_root=tmp_path / "output",
    )
    workbook = load_workbook(result["workbook_path"])
    candidates = workbook["Кандидаты"]
    transcripts = workbook["Транскрипции"]
    assert candidates.cell(2, 15).value == str(video_id)
    assert candidates.cell(2, 15).number_format == "@"
    assert transcripts.cell(2, 2).value == str(video_id)
    assert transcripts.cell(2, 2).number_format == "@"
    assert len(candidates.cell(2, 8).value) <= 300
    assert len(candidates.cell(2, 38).value) <= 500
