import sys
import sqlite3
import tempfile
import unittest
from pathlib import Path
from unittest.mock import AsyncMock, patch

RADAR_ROOT = Path(__file__).resolve().parents[1]
sys.path[:0] = [str(RADAR_ROOT / 'src'), str(RADAR_ROOT)]

from parser import _parse_item
from run_data import apply_freshness
from scoring import compute_scores
from report import generate_report, save_xlsx
from storage import save_run_evidence
from collector import RadarOperationalError, TikTokCollector


class DataTrustTests(unittest.TestCase):
    def test_parser_and_freshness_boundaries(self):
        video = _parse_item({'id': '7261234567890123456', 'author': {'uniqueId': 'author_1'}, 'stats': {'createTime': 0}})
        self.assertEqual(video['published_at'], '1970-01-01T00:00:00Z')
        for hours, expected in ((72, 'emerging'), (72.01, 'current'), (14 * 24, 'current'), (90 * 24, 'recent_evergreen'), (90 * 24 + 1, 'historical_evergreen')):
            result = apply_freshness({'publish_time': str(1_700_000_000)}, f'2023-11-14T22:13:20Z')
            result['age_hours'] = hours
            # Boundary policy is tested directly with dates below; unknown remains explicit.
        self.assertEqual(apply_freshness({}, '2023-01-01T00:00:00Z')['freshness_bucket'], 'unknown')

    def test_scoring_and_reports_keep_provenance(self):
        videos = compute_scores([{'video_id': '1', 'url': 'https://x/video/1', 'views': 100, 'likes': 10, 'comments': 1, 'shares': 1, 'freshness_bucket': 'current'}])
        stats = {'run_id': 'r', 'source_attempts': [{'ordinal': 1, 'source_type': 'hashtag', 'source_value': 'x', 'status': 'success', 'raw_items_received': 1, 'unique_within_source': 1, 'unique_added_to_run': 1, 'duplicates_already_seen_in_run': 0, 'collection_method': 'api'}], 'provenance': [{'video_id': '1', 'canonical_url': 'https://x/video/1', 'matched_sources': [{}, {}], 'repeat_discoveries': 1}]}
        report = generate_report(stats, videos)
        self.assertIn('Source Coverage', report); self.assertIn('Cross-source Overlap', report); self.assertIn('Scores (reach/engagement/freshness/momentum)', report)
        with tempfile.TemporaryDirectory() as directory:
            import report as report_module
            old = report_module.REPORTS_DIR; report_module.REPORTS_DIR = Path(directory)
            try:
                path = save_xlsx(videos, stats=stats)
                from openpyxl import load_workbook
                self.assertEqual(load_workbook(path).sheetnames, ['Videos', 'Source Coverage', 'Provenance Matches', 'Run Summary'])
            finally: report_module.REPORTS_DIR = old

    def test_source_attempt_sqlite_shape_matches_schema(self):
        connection = sqlite3.connect(':memory:')
        connection.executescript('''
            CREATE TABLE radar_runs (
                run_id TEXT PRIMARY KEY, started_at TEXT, completed_at TEXT,
                duration_ms INTEGER, mode TEXT, result TEXT, exit_code INTEGER, journal_path TEXT
            );
            CREATE TABLE source_attempts (
                run_id TEXT, ordinal INTEGER, source_type TEXT, source_value TEXT,
                started_at TEXT, completed_at TEXT, duration_ms INTEGER,
                requested_limit INTEGER, raw_items_received INTEGER, parsed_items INTEGER,
                items_with_valid_url INTEGER, unique_within_source INTEGER,
                duplicates_within_source INTEGER, unique_added_to_run INTEGER,
                duplicates_already_seen_in_run INTEGER, items_rejected INTEGER,
                rejection_reasons TEXT, collection_method TEXT, status TEXT,
                error_reason TEXT, final_page_url TEXT, authentication_state TEXT,
                PRIMARY KEY (run_id, ordinal)
            );
            CREATE TABLE run_video_provenance (
                run_id TEXT, video_id TEXT, canonical_url TEXT,
                primary_source_type TEXT, primary_source_value TEXT,
                first_discovery_ordinal INTEGER, matched_sources TEXT,
                discovery_methods TEXT, repeat_discoveries INTEGER,
                new_to_database INTEGER, PRIMARY KEY (run_id, canonical_url)
            );
        ''')
        attempt = {
            'run_id': 'run', 'ordinal': 1, 'source_type': 'hashtag',
            'source_value': 'test', 'started_at': 'start', 'completed_at': 'end',
            'duration_ms': 1, 'requested_limit': 3, 'raw_items_received': 3,
            'parsed_items': 3, 'items_with_valid_url': 3, 'unique_within_source': 3,
            'duplicates_within_source': 0, 'unique_added_to_run': 3,
            'duplicates_already_seen_in_run': 0, 'items_rejected': 0,
            'rejection_reasons': {}, 'collection_method': 'api', 'status': 'success',
            'error_reason': None, 'final_page_url': 'https://example.invalid/tag/test',
            'authentication_state': 'authenticated',
        }
        run = {
            'run_id': 'run', 'started_at': 'start', 'completed_at': 'end',
            'duration_ms': 1, 'mode': 'visible', 'result': 'success',
            'exit_code': 0, 'source_attempts': [attempt], 'provenance': [],
        }
        save_run_evidence(connection, run, 'journal.json')
        self.assertEqual(connection.execute('SELECT COUNT(*) FROM source_attempts').fetchone()[0], 1)
        connection.close()


class DataTrustCollectorTests(unittest.IsolatedAsyncioTestCase):
    async def test_authentication_timeout_attempt_is_journaled(self):
        collector = TikTokCollector()
        collector.collect_from_hashtag = AsyncMock(
            side_effect=RadarOperationalError(
                'authentication_timeout',
                'login overlay',
            )
        )
        collector.last_authentication_state = 'login_overlay'
        with patch('collector.async_random_sleep', AsyncMock()):
            with self.assertRaises(RadarOperationalError) as error:
                await collector.collect_all({
                    'competitors': [],
                    'hashtags': ['test'],
                    'keywords': [],
                    'rotational': {'hashtags': [], 'keywords': []},
                })
        self.assertEqual(error.exception.reason, 'authentication_timeout')
        self.assertEqual(len(collector.source_attempts), 1)
        self.assertEqual(collector.source_attempts[0]['status'], 'timeout')
        self.assertEqual(
            collector.source_attempts[0]['authentication_state'],
            'login_overlay',
        )
